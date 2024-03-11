#JSE Stock Visualiser
from matplotlib.colors import LinearSegmentedColormap
from openpyxl.styles import NamedStyle, Font
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
from bs4 import BeautifulSoup
from openpyxl import Workbook
import requests
import time

class Stock:
    def __init__(self, ticker, name, price_cent, high, low, close, movement_percent, movement_zar):
        self.ticker = ticker
        self.name = name
        self.price_cent = price_cent
        self.high = high
        self.low = low
        self.close = close
        self.movement_percent = movement_percent
        self.movement_zar = movement_zar
    
    def __str__(self):
        return f"T: {self.ticker}\nN: {self.name}\nP: {self.price_cent}\nH: {self.high}\nL: {self.low}\nC: {self.close}\nMP: {self.movement_percent}\nMR: {self.movement_zar}\n"

# Scraping JSE Share Full Listing data from ABSA web page
url = 'https://www.absa.co.za/indices/share-information/'
response = requests.get(url)

if response.status_code == 200:
    soup = BeautifulSoup(response.content, 'html.parser')
    td_tags = soup.find_all('td')
    
    stocks = []
    stocks_info = []
    for idx, data in enumerate(td_tags):
        text = data.text.strip()
        stocks_info.append(text)
        if len(stocks_info) == 8:
            stocks.append(Stock(*stocks_info))
            stocks_info = []

    tickers = [stock.ticker for stock in stocks]
    movement_percentages = [float(stock.movement_percent) for stock in stocks]

    # Determine the number of rows and columns for the grid layout
    num_cols = 20
    num_rows = len(tickers) // num_cols + (1 if len(tickers) % num_cols != 0 else 0) #Add an extra row if there are remaining objects

    # Custom colormap with dark red and dark green
    colors = [(0.5, 0, 0), (0, 0.5, 0)]
    cmap = LinearSegmentedColormap.from_list('CustomDarkRedGreen', colors)

    # Get the most extreme percentage
    absolute_percentages = [abs(mp) for mp in movement_percentages]
    ymax = max(absolute_percentages)
    # Cap maximum y-axis value at 15
    if ymax > 15:
        ymax = 15
    ymin = ymax*-1
    norm = plt.Normalize(ymin, ymax)

    plt.figure(figsize=(15, 8))
    plt.subplots_adjust(right=0.95, left=0.05, top=0.95, bottom=0.05)

    # Plot colored rectangles for each stock cell
    for i, (ticker, movement_percent) in enumerate(zip(tickers, movement_percentages)):
        row = i // num_cols
        col = i % num_cols
        color = cmap(norm(movement_percent))
        plt.fill([col, col+1, col+1, col], [row, row, row+1, row+1], color=color)
        plt.text(col + 0.5, row + 0.5, ticker, color='white', ha='center', va='bottom', fontsize=10)
        plt.text(col + 0.5, row + 0.5, f"{movement_percent}%", color='white', ha='center', va='top', fontsize=8)

    plt.title('Market Carpet')
    plt.gca().invert_yaxis()
    plt.tight_layout()
    plt.axis('off')
    
    # Creating a new excel workbook
    wb = Workbook()
    ws = wb.active
    header_style = NamedStyle(name='header')
    header_style.font = Font(bold=True, size=12)
    header_row = ["Ticker", "Name", "Price (c)", "High", "Low", "Close", "Movement (%)", "Movement (R)"]
    ws.append(header_row)
    ws.append([None, None, None, None, None, None, None, None])
    ws.append([None, None, None, None, None, None, None, None])

    for cell in ws[1]:
        cell.style = header_style
    
    # Writing stock data to the worksheet
    for stock in stocks:
        stock_data = [stock.ticker, stock.name]
        for value in [stock.price_cent, stock.high, stock.low, stock.close, stock.movement_percent, stock.movement_zar]:
            if value is not None:
                try:                  
                    parsed_value = int(value)
                except ValueError:
                    try:
                        parsed_value = float(value)
                    except ValueError:
                        parsed_value = value
            else:
                parsed_value = None
            stock_data.append(parsed_value)
        ws.append(stock_data)

    # Generating filename with current Unix time
    unix_time = int(time.time())
    filename = f"JSE_SFL_{unix_time}.xlsx"

    # Save plot to a PNG file
    fig_name = f"MarketCarpet_{unix_time}.png"
    plt.savefig(fig_name, bbox_inches='tight')

    # Insert the PNG image
    img = Image(fig_name)
    ws.add_image(img, 'I1')

    # Define column widths
    column_widths = [8.11, 27, 8.5, 8.11, 8.11, 8.11, 14, 14]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # Create a border object with the desired style and colour
    border_style = 'thin'
    border_color = 'DDDDDD' 
    border = Border(left=Side(style=border_style, color=border_color),
                    right=Side(style=border_style, color=border_color),
                    top=Side(style=border_style, color=border_color),
                    bottom=Side(style=border_style, color=border_color))

    for row in ws.iter_rows(min_row=4, min_col=1, max_col=8):
        # Read the value of the cell in the 7th column (movement_percent)
        value = row[6].value
        # Formatting row fill colours based on value
        if value is not None and isinstance(value, (int, float)):
            if value < 0:
                fill = PatternFill(start_color='FFBDBD', end_color='FFBDBD', fill_type='solid')  # Light red fill colour
                for idx in range(8):
                    row[idx].fill = fill
                    row[idx].border = border
            elif value > 0:
                fill = PatternFill(start_color='BDFFBD', end_color='BDFFBD', fill_type='solid')  # Light green fill colour
                for idx in range(8):
                    row[idx].fill = fill
                    row[idx].border = border

    # Save the workbook
    wb.save(filename)
    
    print(f"Stock data successfully saved to '{filename}'")
else:
    print('Failed to retrieve the webpage')




